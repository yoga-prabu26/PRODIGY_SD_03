"""
contact_manager.py — Business logic / controller layer for ContactVault.
"""

import csv
import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional, Callable

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from models import Contact, Category, AppStats, ActivityType
from database import Database

EXPORTS_DIR = Path(__file__).parent / "exports"
BACKUPS_DIR = Path(__file__).parent / "backups"
AVATARS_DIR = Path(__file__).parent / "assets" / "avatars"

for _d in (EXPORTS_DIR, BACKUPS_DIR, AVATARS_DIR):
    _d.mkdir(parents=True, exist_ok=True)


class ContactManager:
    """Controller: bridges UI actions with the database and I/O."""

    def __init__(self):
        self.db = Database()
        self._listeners: list[Callable] = []   # for observer pattern

    # ── observer ──────────────────────────────────────────────────
    def subscribe(self, fn: Callable):
        self._listeners.append(fn)

    def _notify(self, event: str, data=None):
        for fn in self._listeners:
            try:
                fn(event, data)
            except Exception:
                pass

    # ── contacts ─────────────────────────────────────────────────
    def add_contact(self, c: Contact) -> tuple:
        errors = c.validate()
        if errors:
            return False, errors

        dupes = self.db.find_duplicates(c.full_name, c.mobile)
        if dupes:
            return "duplicate", dupes

        cid = self.db.add_contact(c)
        c.id = cid
        self._notify("contact_added", c)
        return True, cid

    def update_contact(self, c: Contact) -> tuple:
        errors = c.validate()
        if errors:
            return False, errors

        dupes = self.db.find_duplicates(c.full_name, c.mobile, exclude_id=c.id)
        if dupes:
            return "duplicate", dupes

        self.db.update_contact(c)
        self._notify("contact_updated", c)
        return True, c.id

    def delete_contact(self, contact_id: int, name: str = "") -> bool:
        ok = self.db.delete_contact(contact_id, name)
        if ok:
            self._notify("contact_deleted", contact_id)
        return ok

    def get_contact(self, cid: int) -> Optional[Contact]:
        c = self.db.get_contact(cid)
        if c:
            self.db._log(ActivityType.VIEWED, c.full_name)
        return c

    def get_all_contacts(self, sort_by: str = "full_name",
                          order: str = "ASC") -> list:
        return self.db.get_all_contacts(sort_by, order)

    def search(self, query: str) -> list:
        if not query.strip():
            return self.db.get_all_contacts()
        return self.db.search_contacts(query)

    def filter_by_category(self, category: Category) -> list:
        return self.db.get_by_category(category)

    def get_favorites(self) -> list:
        return self.db.get_favorites()

    def toggle_favorite(self, contact_id: int) -> bool:
        state = self.db.toggle_favorite(contact_id)
        self._notify("favorite_toggled", (contact_id, state))
        return state

    def get_stats(self) -> AppStats:
        return self.db.get_stats()

    def get_activity_log(self, limit: int = 50) -> list:
        return self.db.get_activity_log(limit)

    # ── avatar ────────────────────────────────────────────────────
    def save_avatar(self, contact_id: int, src_path: str) -> str:
        ext = Path(src_path).suffix
        dest = AVATARS_DIR / f"contact_{contact_id}{ext}"
        shutil.copy2(src_path, dest)
        return str(dest)

    # ── CSV export ────────────────────────────────────────────────
    def export_csv(self, contacts: list = None) -> Path:
        contacts = contacts or self.db.get_all_contacts()
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = EXPORTS_DIR / f"contacts_export_{ts}.csv"
        fieldnames = ["full_name", "mobile", "email", "address",
                      "company", "job_title", "category", "tags",
                      "is_favorite", "notes", "created_at"]
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for c in contacts:
                row = c.to_dict()
                writer.writerow({k: row[k] for k in fieldnames})
        self.db._log(ActivityType.EXPORTED, f"{len(contacts)} contacts", "CSV")
        return path

    # ── Excel export ──────────────────────────────────────────────
    def export_excel(self, contacts: list = None) -> Path:
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl not installed.")
        contacts = contacts or self.db.get_all_contacts()
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = EXPORTS_DIR / f"contacts_export_{ts}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contacts"

        headers = ["Name", "Mobile", "Email", "Address", "Company",
                   "Job Title", "Category", "Tags", "Favorite",
                   "Notes", "Created At"]
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(bold=True, color="FFFFFF", name="Calibri")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, c in enumerate(contacts, 2):
            d = c.to_dict()
            ws.append([
                d["full_name"], d["mobile"], d["email"], d["address"],
                d["company"], d["job_title"], d["category"], d["tags"],
                "Yes" if d["is_favorite"] else "No",
                d["notes"], d["created_at"]
            ])
            if row_idx % 2 == 0:
                for col in range(1, 12):
                    ws.cell(row_idx, col).fill = PatternFill(
                        "solid", fgColor="F0F4FF")

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

        wb.save(path)
        self.db._log(ActivityType.EXPORTED, f"{len(contacts)} contacts", "Excel")
        return path

    # ── CSV import ────────────────────────────────────────────────
    def import_csv(self, file_path: str) -> tuple:
        added, skipped, errors = 0, 0, []
        try:
            with open(file_path, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for i, row in enumerate(reader, 1):
                    try:
                        c = Contact(
                            full_name=row.get("full_name", row.get("Name", "")).strip(),
                            mobile=row.get("mobile", row.get("Mobile", "")).strip(),
                            email=row.get("email", row.get("Email", "")).strip(),
                            address=row.get("address", row.get("Address", "")).strip(),
                            company=row.get("company", row.get("Company", "")).strip(),
                            job_title=row.get("job_title", row.get("Job Title", "")).strip(),
                            notes=row.get("notes", row.get("Notes", "")).strip(),
                            category=Category.from_str(
                                row.get("category", row.get("Category", "Personal"))
                            ),
                            tags=row.get("tags", row.get("Tags", "")).strip(),
                        )
                        result, _ = self.add_contact(c)
                        if result is True:
                            added += 1
                        else:
                            skipped += 1
                    except Exception as e:
                        errors.append(f"Row {i}: {e}")
        except Exception as e:
            return 0, 0, [str(e)]

        self.db._log(ActivityType.IMPORTED,
                     f"{added} contacts", f"from {file_path}")
        self._notify("contacts_imported", added)
        return added, skipped, errors

    # ── backup / restore ──────────────────────────────────────────
    def backup(self) -> Path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = BACKUPS_DIR / f"contacts_backup_{ts}.db"
        self.db.backup(dest)
        return dest

    def restore(self, src: str) -> bool:
        ok = self.db.restore(Path(src))
        if ok:
            self._notify("db_restored", None)
        return ok
